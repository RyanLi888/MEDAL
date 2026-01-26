"""
Label Correction Analysis Script
Only runs label correction part and generates detailed analysis documents and feature distribution plots.

This script:
1. Loads dataset and injects label noise
2. Extracts features using pre-trained backbone
3. Runs Hybrid Court label correction (CL + AUM + KNN)
4. Generates detailed per-sample analysis document
5. Creates feature distribution plots for clean, noisy, and corrected data
"""
import sys
import os
from pathlib import Path
# Ensure project root is on sys.path so `MoudleCode.*` imports work when invoked directly
PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.append(str(PROJECT_ROOT))

import torch
import numpy as np
import pandas as pd
import argparse
from datetime import datetime
import json
import matplotlib.pyplot as plt
import matplotlib as mpl
from matplotlib import font_manager
import seaborn as sns
from sklearn.manifold import TSNE
from sklearn.decomposition import PCA

from MoudleCode.utils.config import config
from MoudleCode.utils.helpers import (
    set_seed, setup_logger, inject_label_noise
)
from MoudleCode.preprocessing.pcap_parser import load_dataset
from MoudleCode.feature_extraction.backbone import MicroBiMambaBackbone, build_backbone
from MoudleCode.label_correction.hybrid_court import HybridCourt

import logging

# 尝试导入Excel写入库
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: openpyxl未安装，将只生成CSV文件。安装命令: pip install openpyxl")

# 简化的预处理文件检测和加载
def check_preprocessed_exists(split='train'):
    """检查预处理文件是否存在"""
    preprocessed_dir = Path(config.OUTPUT_ROOT) / 'preprocessed'
    if not preprocessed_dir.exists():
        return False
    
    X_file = preprocessed_dir / f'{split}_X.npy'
    y_file = preprocessed_dir / f'{split}_y.npy'
    files_file = preprocessed_dir / f'{split}_files.npy'
    
    return X_file.exists() and y_file.exists() and files_file.exists()

def load_preprocessed(split='train'):
    """加载预处理文件"""
    preprocessed_dir = Path(config.OUTPUT_ROOT) / 'preprocessed'
    
    X = np.load(preprocessed_dir / f'{split}_X.npy')
    y = np.load(preprocessed_dir / f'{split}_y.npy')
    files = np.load(preprocessed_dir / f'{split}_files.npy', allow_pickle=True)
    
    return X, y, files

logger = None  # Will be initialized in main()


def _configure_matplotlib_chinese_fonts():
    """Configure matplotlib to use Chinese fonts and suppress warnings if unavailable."""
    import warnings
    # Suppress font warnings globally if no Chinese font is available
    warnings.filterwarnings('ignore', category=UserWarning, 
                          message='.*Glyph.*missing from font.*')
    
    try:
        preferred = [
            'Noto Sans CJK SC',
            'Noto Sans CJK TC',
            'Noto Sans CJK JP',
            'Noto Sans SC',
            'Noto Sans TC',
            'Source Han Sans SC',
            'Source Han Sans CN',
            'WenQuanYi Zen Hei',
            'WenQuanYi Micro Hei',
            'Microsoft YaHei',
            'SimHei'
        ]
        available = {f.name for f in font_manager.fontManager.ttflist}
        chosen = None
        for name in preferred:
            if name in available:
                chosen = name
                break

        if chosen is not None:
            current = mpl.rcParams.get('font.sans-serif', [])
            if not isinstance(current, list):
                current = [current]
            mpl.rcParams['font.family'] = 'sans-serif'
            mpl.rcParams['font.sans-serif'] = [chosen] + [f for f in current if f != chosen]

        mpl.rcParams['axes.unicode_minus'] = False
    except Exception:
        # If font configuration fails, warnings are already suppressed above
        pass


_configure_matplotlib_chinese_fonts()


def extract_features_with_backbone(backbone, X_data, config, logger):
    """
    Extract features using pre-trained backbone
    
    Args:
        backbone: Pre-trained MicroBiMambaBackbone
        X_data: (N, L, D) input sequences
        config: configuration object
        logger: logger
        
    Returns:
        features: (N, d) extracted features
    """
    logger.info("Extracting features using backbone...")
    
    backbone.freeze()
    backbone.eval()
    backbone.to(config.DEVICE)
    
    features_list = []
    batch_size = 64
    X_tensor = torch.FloatTensor(X_data).to(config.DEVICE)
    
    total_batches = (len(X_tensor) + batch_size - 1) // batch_size
    
    with torch.no_grad():
        for i in range(0, len(X_tensor), batch_size):
            batch_idx = i // batch_size + 1
            X_batch = X_tensor[i:i+batch_size]
            z_batch = backbone(X_batch, return_sequence=False)
            features_list.append(z_batch.cpu().numpy())
            
            if batch_idx % 10 == 0 or batch_idx == total_batches:
                progress = batch_idx / total_batches * 100
                logger.info(f"  Feature extraction progress: {batch_idx}/{total_batches} batches ({progress:.1f}%)")
    
    features = np.concatenate(features_list, axis=0)
    logger.info(f"✓ Feature extraction complete: {features.shape}")
    
    return features


def run_hybrid_court_with_detailed_tracking(features, noisy_labels, config, logger, y_true=None):
    """
    Run Hybrid Court (CL+AUM+KNN Two-Phase, No-Drop) with detailed tracking
    
    Args:
        features: 特征矩阵
        noisy_labels: 噪声标签
        config: 配置对象
        logger: 日志器
        y_true: 真实标签 (可选，用于详细统计)
    
    Returns:
        results: dict containing all intermediate and final results
    """
    logger.info("")
    logger.info("="*70)
    logger.info("Running Hybrid Court (CL+AUM+KNN)")
    logger.info("="*70)
    
    hybrid_court = HybridCourt(config)
    n_samples = len(noisy_labels)

    clean_labels, action_mask, confidence, correction_weight, aum_scores, neighbor_consistency, pred_probs = hybrid_court.correct_labels(
        features=features,
        noisy_labels=noisy_labels,
        device=str(config.DEVICE),
        y_true=y_true,
    )
    
    # 从各子模块中读取缓存的中间结果
    suspected_noise = getattr(hybrid_court.cl, "last_suspected_noise", None)
    pred_labels = getattr(hybrid_court.cl, "last_pred_labels", None)
    neighbor_labels = getattr(hybrid_court.knn, "last_neighbor_labels", None)
    action_names = ['Keep', 'Flip', 'Drop', 'Reweight']
    tier_info = [action_names[int(a)] if int(a) < len(action_names) else '' for a in np.asarray(action_mask).tolist()]
    
    # 获取迭代CL和锚点KNN结果
    iter_pred_probs = getattr(hybrid_court, "iter_pred_probs_all", None)
    anchor_votes = getattr(hybrid_court, "anchor_votes_all", None)
    anchor_consistency = getattr(hybrid_court, "anchor_consistency_all", None)

    # Phase1 矫正后重算的 Stage2 指标（供二阶段策略使用）
    stage2_aum_scores = getattr(hybrid_court, "stage2_aum_scores_all", None)
    stage2_knn_neighbor_labels = getattr(hybrid_court, "stage2_knn_neighbor_labels_all", None)
    stage2_knn_neighbor_consistency = getattr(hybrid_court, "stage2_knn_neighbor_consistency_all", None)
    stage2_cl_suspected_noise = getattr(hybrid_court, "stage2_cl_suspected_noise_all", None)
    stage2_cl_pred_labels = getattr(hybrid_court, "stage2_cl_pred_labels_all", None)
    stage2_cl_pred_probs = getattr(hybrid_court, "stage2_cl_pred_probs_all", None)
    phase2_actions = getattr(hybrid_court, "phase2_action_all", None)
    
    # 生成决策原因
    decision_reasons = []
    for i in range(n_samples):
        tier = tier_info[i] if i < len(tier_info) else ''
        
        # 获取迭代CL和锚点KNN信息
        # iter_pred_probs 对应 “Phase1 矫正后标签” 的 CL 概率，因此当前标签用 clean_labels
        iter_cl_current = float(iter_pred_probs[i, int(clean_labels[i])]) if iter_pred_probs is not None else None
        iter_cl_target = float(iter_pred_probs[i, 1 - int(clean_labels[i])]) if iter_pred_probs is not None else None
        anchor_vote = int(anchor_votes[i]) if anchor_votes is not None else None
        anchor_cons = float(anchor_consistency[i]) if anchor_consistency is not None else None
        
        reason = _generate_decision_reason(
            tier,
            int(noisy_labels[i]),
            int(clean_labels[i]),
            int(action_mask[i]),
            bool(suspected_noise[i]) if suspected_noise is not None else False,
            int(neighbor_labels[i]) if neighbor_labels is not None else -1,
            float(neighbor_consistency[i]),
            float(pred_probs[i, 0]),
            float(pred_probs[i, 1]),
            float(aum_scores[i]) if aum_scores is not None else 0.0,
            iter_cl_current,
            iter_cl_target,
            anchor_vote,
            anchor_cons
        )
        decision_reasons.append(reason)
    
    # Compile results
    results = {
        'features': features,
        'noisy_labels': noisy_labels,
        'clean_labels': clean_labels,
        'action_mask': action_mask,
        'confidence': confidence,
        'correction_weight': correction_weight,
        # CL results
        'cl_suspected_noise': suspected_noise,
        'cl_pred_labels': pred_labels,
        'cl_pred_probs': pred_probs,
        'aum_scores': aum_scores,
        # KNN results
        'knn_neighbor_labels': neighbor_labels,
        'knn_neighbor_consistency': neighbor_consistency,
        # Iterative CL results
        'iter_cl_probs': iter_pred_probs,
        # Anchor KNN results
        'anchor_votes': anchor_votes,
        'anchor_consistency': anchor_consistency,
        # Stage2 metrics (recomputed on phase1 corrected labels)
        'stage2_aum_scores': stage2_aum_scores,
        'stage2_knn_neighbor_labels': stage2_knn_neighbor_labels,
        'stage2_knn_neighbor_consistency': stage2_knn_neighbor_consistency,
        'stage2_cl_suspected_noise': stage2_cl_suspected_noise,
        'stage2_cl_pred_labels': stage2_cl_pred_labels,
        'stage2_cl_pred_probs': stage2_cl_pred_probs,
        'phase2_actions': phase2_actions,
        # Decision tracking
        'decision_reasons': decision_reasons,
        'tier_info': tier_info
    }
    
    # 详细的Tier统计
    _log_tier_statistics(results, y_true, logger)
    
    return results


def _generate_decision_reason(tier, noisy_label, clean_label, action, is_suspected,
                               knn_label, knn_cons, cl_prob_0, cl_prob_1, aum,
                               iter_cl_current=None, iter_cl_target=None,
                               anchor_vote=None, anchor_cons=None):
    """生成详细的决策原因说明"""
    current_label_name = '正常' if noisy_label == 0 else '恶意'
    target_label_name = '恶意' if noisy_label == 0 else '正常'
    new_label_name = '正常' if clean_label == 0 else '恶意'
    
    # 原始CL置信度
    orig_cl_current = cl_prob_0 if noisy_label == 0 else cl_prob_1
    orig_cl_target = cl_prob_1 if noisy_label == 0 else cl_prob_0
    
    # KNN投票结果
    knn_vote_name = '正常' if knn_label == 0 else '恶意'
    knn_support = '支持' if knn_label == noisy_label else '反对'
    
    if action == 1:
        cl_diff = orig_cl_target - orig_cl_current
        knn_vote_name = '正常' if knn_label == 0 else '恶意'
        return (
            f"[Flip] CL差值={cl_diff:.3f} | AUM={aum:.3f} | "
            f"KNN投票={knn_vote_name}(cons={knn_cons:.3f}) → 翻转{current_label_name}→{new_label_name}"
        )
    if action == 2:
        return (
            f"[Drop] CL={orig_cl_current:.3f} | AUM={aum:.3f} | "
            f"KNN(cons={knn_cons:.3f}) → 丢弃{current_label_name}"
        )
    if action == 3:
        return (
            f"[Reweight] CL={orig_cl_current:.3f} | AUM={aum:.3f} | "
            f"KNN(cons={knn_cons:.3f}) → 重加权{current_label_name}"
        )
    return (
        f"[Keep] CL={orig_cl_current:.3f} | AUM={aum:.3f} | "
        f"KNN(cons={knn_cons:.3f}) → 保持{current_label_name}"
    )
    
    # Tier 4: Reweight - 不确定样本
    if 'Tier 4' in tier or 'Reweight' in tier:
        reasons = []
        if orig_cl_current < 0.55:
            reasons.append(f"原CL低({orig_cl_current:.3f}<0.55)")
        if knn_cons < 0.55:
            reasons.append(f"KNN弱({knn_cons:.3f}<0.55)")
        if not is_suspected and knn_label == noisy_label:
            reasons.append("CL不怀疑+KNN支持")
        
        if not reasons:
            reasons.append("未达Core/Flip/Keep标准")
        
        return f"[Phase2-Reweight] {' + '.join(reasons)} → 降权保持{current_label_name}(w=0.5)"
    
    # Tier 5: Rescued - Phase 3拯救
    if 'Tier 5' in tier or 'Rescued' in tier:
        if anchor_cons is not None and anchor_vote is not None:
            anchor_vote_name = '正常' if anchor_vote == 0 else '恶意'
            if 'Keep' in tier or '5a' in tier:
                return (f"[Phase3-Rescued-Keep] iter_T={iter_cl_target:.3f}≥0.60 + "
                       f"anchor强支持当前({anchor_vote_name},{anchor_cons:.3f}≥0.60) → "
                       f"拯救保持{current_label_name}(w=0.85)")
            else:
                return (f"[Phase3-Rescued-Flip] iter_T={iter_cl_target:.3f}≥0.60 + "
                       f"anchor强反对当前({anchor_vote_name},{anchor_cons:.3f}≥0.60) → "
                       f"拯救翻转{current_label_name}→{new_label_name}(w=0.75)")
        else:
            return f"[Phase3-Rescued] 锚点拯救 → {new_label_name}"
    
    # Dropped - 丢弃
    if 'Dropped' in tier:
        drop_reasons = []
        if iter_cl_current is not None:
            if iter_cl_current < 0.48:
                drop_reasons.append(f"iter_CL极低({iter_cl_current:.3f}<0.48)")
            elif iter_cl_current < 0.55 and anchor_cons is not None and anchor_cons < 0.55:
                drop_reasons.append(f"iter_CL低({iter_cl_current:.3f}<0.55) + anchor低({anchor_cons:.3f}<0.55)")
        
        if not drop_reasons:
            drop_reasons.append("质量过低")
        
        return f"[Dropped] {' 或 '.join(drop_reasons)} → 丢弃(w=0.0)"
    
    # 未知Tier
    else:
        action_names = ['保持', '翻转', '丢弃', '重加权']
        action_name = action_names[action] if action < len(action_names) else '未知'
        return f"[未分类] tier={tier}, action={action_name}, 原CL={orig_cl_current:.3f}, KNN={knn_cons:.3f}"


def _log_tier_statistics(results, y_true, logger):
    """输出详细的Tier统计"""
    tier_info = results.get('tier_info', [])
    clean_labels = results['clean_labels']
    correction_weight = results['correction_weight']
    n_samples = len(clean_labels)
    
    # 统计各Tier
    tier_counts = {}
    tier_correct = {}
    tier_weights = {}
    
    for i, tier in enumerate(tier_info):
        if tier not in tier_counts:
            tier_counts[tier] = 0
            tier_correct[tier] = 0
            tier_weights[tier] = correction_weight[i]
        tier_counts[tier] += 1
        if y_true is not None and clean_labels[i] == y_true[i]:
            tier_correct[tier] += 1
    
    # 两阶段策略不使用Tier统计，删除无用输出


def generate_sample_analysis_document(results, y_true, noise_mask, save_path, logger, noise_pct=None):
    """
    Generate detailed per-sample analysis document in strategy-grouped format
    
    Args:
        results: dict from run_hybrid_court_with_detailed_tracking
        y_true: (N,) ground truth labels
        noise_mask: (N,) boolean array indicating which labels were flipped to create noise
        save_path: path to save the document
        logger: logger
        noise_pct: noise percentage (for filename)
    """
    logger.info("")
    logger.info("="*70)
    logger.info("生成样本级分析文档（策略分组格式）")
    logger.info("="*70)
    
    def _classify_error_type(is_noise, action, true_label, noisy_label, corrected_label):
        """分类错误类型"""
        if corrected_label == true_label:
            return '正确'
        
        if is_noise:
            # 噪声样本
            if action == 1:  # 翻转
                return '翻转错误-噪声未矫正'
            else:
                return '保持错误-噪声未检测'
        else:
            # 干净样本
            if action == 1:  # 翻转
                return '翻转错误-误杀干净样本'
            else:
                return '保持错误-不应发生'
    
    n_samples = len(y_true)
    tier_info = results.get('tier_info', [''] * n_samples)
    
    # 获取迭代CL和锚点KNN的结果（如果存在）
    iter_cl_probs = results.get('iter_cl_probs', None)
    anchor_votes = results.get('anchor_votes', None)
    anchor_consistency = results.get('anchor_consistency', None)

    # Phase1 矫正后重算的 Stage2 指标（如果存在）
    stage2_aum_scores = results.get('stage2_aum_scores', None)
    stage2_knn_neighbor_labels = results.get('stage2_knn_neighbor_labels', None)
    stage2_knn_neighbor_consistency = results.get('stage2_knn_neighbor_consistency', None)
    stage2_cl_suspected_noise = results.get('stage2_cl_suspected_noise', None)
    stage2_cl_pred_labels = results.get('stage2_cl_pred_labels', None)
    stage2_cl_pred_probs = results.get('stage2_cl_pred_probs', None)
    phase2_actions = results.get('phase2_actions', None)
    
    # Prepare data for DataFrame
    data = []
    
    for i in range(n_samples):
        # Basic info
        sample_id = i
        true_label = int(y_true[i])
        true_label_name = "正常" if true_label == 0 else "恶意"
        is_noise = bool(noise_mask[i])
        noisy_label = int(results['noisy_labels'][i])
        noisy_label_name = "正常" if noisy_label == 0 else "恶意"
        
        # CL (Confident Learning) results
        cl_suspected_noise = bool(results['cl_suspected_noise'][i])
        cl_pred_label = int(results['cl_pred_labels'][i])
        cl_pred_label_name = "正常" if cl_pred_label == 0 else "恶意"
        cl_pred_prob_benign = float(results['cl_pred_probs'][i, 0])
        cl_pred_prob_malicious = float(results['cl_pred_probs'][i, 1])
        
        aum_scores = results.get('aum_scores', None)
        aum_score = float(aum_scores[i]) if aum_scores is not None and hasattr(aum_scores, '__getitem__') and len(aum_scores) > i else 0.0
        
        # KNN (Semantic Voting) results
        knn_neighbor_label = int(results['knn_neighbor_labels'][i])
        knn_neighbor_label_name = "正常" if knn_neighbor_label == 0 else "恶意"
        knn_consistency = float(results['knn_neighbor_consistency'][i])
        
        corrected_label = int(results['clean_labels'][i])

        # Iterative CL results (if available) - computed on corrected labels
        if iter_cl_probs is not None:
            iter_cl_current = float(iter_cl_probs[i, corrected_label])
            iter_cl_target = float(iter_cl_probs[i, 1 - corrected_label])
        else:
            iter_cl_current = None
            iter_cl_target = None

        # Stage2 metrics (if available)
        if stage2_aum_scores is not None and hasattr(stage2_aum_scores, '__getitem__') and len(stage2_aum_scores) > i:
            stage2_aum = float(stage2_aum_scores[i])
        else:
            stage2_aum = None

        if stage2_knn_neighbor_labels is not None and hasattr(stage2_knn_neighbor_labels, '__getitem__') and len(stage2_knn_neighbor_labels) > i:
            stage2_knn_label = int(stage2_knn_neighbor_labels[i])
            stage2_knn_label_name = "正常" if stage2_knn_label == 0 else "恶意"
        else:
            stage2_knn_label = None
            stage2_knn_label_name = None

        if stage2_knn_neighbor_consistency is not None and hasattr(stage2_knn_neighbor_consistency, '__getitem__') and len(stage2_knn_neighbor_consistency) > i:
            stage2_knn_cons = float(stage2_knn_neighbor_consistency[i])
        else:
            stage2_knn_cons = None

        if stage2_cl_suspected_noise is not None and hasattr(stage2_cl_suspected_noise, '__getitem__') and len(stage2_cl_suspected_noise) > i:
            stage2_cl_noise_flag = bool(stage2_cl_suspected_noise[i])
        else:
            stage2_cl_noise_flag = None

        if stage2_cl_pred_labels is not None and hasattr(stage2_cl_pred_labels, '__getitem__') and len(stage2_cl_pred_labels) > i:
            stage2_cl_pred_label = int(stage2_cl_pred_labels[i])
            stage2_cl_pred_label_name = "正常" if stage2_cl_pred_label == 0 else "恶意"
        else:
            stage2_cl_pred_label = None
            stage2_cl_pred_label_name = None

        if stage2_cl_pred_probs is not None and hasattr(stage2_cl_pred_probs, '__getitem__') and len(stage2_cl_pred_probs) > i:
            stage2_cl_pred_prob_benign = float(stage2_cl_pred_probs[i, 0])
            stage2_cl_pred_prob_malicious = float(stage2_cl_pred_probs[i, 1])
        else:
            stage2_cl_pred_prob_benign = None
            stage2_cl_pred_prob_malicious = None

        if phase2_actions is not None and hasattr(phase2_actions, '__getitem__') and len(phase2_actions) > i:
            phase2_action = str(phase2_actions[i])
        else:
            phase2_action = ''
        
        # Anchor KNN results (if available)
        if anchor_votes is not None:
            anchor_vote = int(anchor_votes[i])
            anchor_vote_name = "正常" if anchor_vote == 0 else "恶意"
            anchor_cons = float(anchor_consistency[i])
        else:
            anchor_vote = None
            anchor_vote_name = None
            anchor_cons = None
        
        # Final correction decision
        action = int(results['action_mask'][i])
        action_names = ['保持', '翻转', '丢弃', '重加权']
        action_name = action_names[action]
        corrected_label_name = "正常" if corrected_label == 0 else "恶意"
        confidence = float(results['confidence'][i])
        correction_weight = float(results['correction_weight'][i])
        decision_reason = results['decision_reasons'][i]
        tier = tier_info[i] if i < len(tier_info) else ''
        
        # Correction correctness (compared to ground truth)
        if action == 2:  # Dropped
            correction_correct = "不适用(已丢弃)"
        else:
            correction_correct = "正确" if corrected_label == true_label else "错误"
        
        # Compile row with ALL detailed metrics
        row = {
            # 基本信息
            '样本ID': sample_id,
            '真实标签': true_label_name,
            '真实标签值': true_label,
            '是否噪声': '是' if is_noise else '否',
            '噪声标签': noisy_label_name,
            '噪声标签值': noisy_label,
            
            # CL (Confident Learning) 详细结果
            'CL疑似噪声': '是' if cl_suspected_noise else '否',
            'CL预测标签': cl_pred_label_name,
            'CL预测标签值': cl_pred_label,
            'CL正常概率': cl_pred_prob_benign,
            'CL恶意概率': cl_pred_prob_malicious,
            'CL当前标签置信度': cl_pred_prob_benign if noisy_label == 0 else cl_pred_prob_malicious,
            'CL目标标签置信度': cl_pred_prob_malicious if noisy_label == 0 else cl_pred_prob_benign,
            'AUM分数': aum_score,
            
            # KNN (Semantic Voting) 详细结果
            'KNN邻居标签': knn_neighbor_label_name,
            'KNN邻居标签值': knn_neighbor_label,
            'KNN一致性': knn_consistency,
            'KNN是否支持当前标签': '是' if knn_neighbor_label == noisy_label else '否',
            'KNN一致性等级': 'High' if knn_consistency >= 0.7 else ('Medium' if knn_consistency >= 0.5 else 'Low'),
        }

        # Add Stage2 metrics if available
        if stage2_aum is not None:
            row['stage2_AUM分数'] = stage2_aum
        if stage2_knn_label is not None:
            row['stage2_KNN邻居标签'] = stage2_knn_label_name
            row['stage2_KNN邻居标签值'] = stage2_knn_label
        if stage2_knn_cons is not None:
            row['stage2_KNN一致性'] = stage2_knn_cons
            # stage2_KNN是否支持当前标签：应该基于Phase1矫正后的标签判断
            # 这里先用corrected_label，后续在阶段2sheet中会重新计算
            row['stage2_KNN是否支持当前标签'] = '是' if stage2_knn_label == corrected_label else '否'

        if stage2_cl_noise_flag is not None:
            # numeric 0/1 for easier pandas operations
            row['stage2_CL疑似噪声'] = 1 if stage2_cl_noise_flag else 0
        
        # Add stage2_CL预测标签 if available
        if stage2_cl_pred_label is not None:
            row['stage2_CL预测标签'] = stage2_cl_pred_label_name
            row['stage2_CL预测标签值'] = stage2_cl_pred_label
            if stage2_cl_pred_prob_benign is not None:
                # 计算基于Phase1矫正后标签的CL置信度
                phase1_corrected_label = corrected_label  # 这里corrected_label就是Phase1矫正后的标签
                row['stage2_CL当前标签置信度'] = stage2_cl_pred_prob_benign if phase1_corrected_label == 0 else stage2_cl_pred_prob_malicious
                row['stage2_CL目标标签置信度'] = stage2_cl_pred_prob_malicious if phase1_corrected_label == 0 else stage2_cl_pred_prob_benign
        
        # Add iterative CL columns if available
        if iter_cl_current is not None:
            row['iter_CL当前标签置信度'] = iter_cl_current
            row['iter_CL目标标签置信度'] = iter_cl_target
            row['iter_CL置信度差值'] = abs(iter_cl_current - iter_cl_target)
        
        # Add anchor KNN columns if available
        if anchor_vote is not None:
            row['anchor_KNN投票'] = anchor_vote_name
            row['anchor_KNN投票值'] = anchor_vote
            row['anchor_KNN一致性'] = anchor_cons
            row['anchor_KNN是否支持当前'] = '是' if anchor_vote == corrected_label else '否'
        
        # 最终决策详细信息
        row.update({
            '矫正动作': action_name,
            '矫正动作值': action,
            '矫正后标签': corrected_label_name,
            '矫正后标签值': corrected_label,
            'Phase2_Action': phase2_action,
            '系统置信度': confidence,
            '样本权值': correction_weight,
            'Tier分级': tier,
            'Tier阶段': 'Dropped' if action == 2 else ('Phase 2' if action in (1, 3) else 'Phase 1'),
            '决策理由': decision_reason,
            
            # 矫正结果评估
            '矫正是否正确': correction_correct,
            '是否翻转': '是' if action == 1 else '否',
            '是否丢弃': '是' if action == 2 else '否',
            '标签是否改变': '是' if corrected_label != noisy_label else '否',
            
            # 错误类型分析
            '错误类型': _classify_error_type(
                is_noise, action, true_label, noisy_label, corrected_label
            ) if action != 2 else '已丢弃'
        })
        
        data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to CSV (基础版本，所有数据在一个文件)
    # 新策略版本化：确保文件名包含噪声率与策略标识，避免覆盖旧结果
    if noise_pct is not None:
        base_name = save_path.replace('.txt', '').replace('.log', '')
        if f'_{noise_pct}pct' not in base_name:
            base_name = f"{base_name}_{noise_pct}pct"
        if '_cl_aum_knn' not in base_name:
            base_name = f"{base_name}_cl_aum_knn"
        csv_path = f"{base_name}.csv"
    else:
        csv_path = save_path.replace('.txt', '.csv').replace('.log', '.csv')
    
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    logger.info(f"✓ CSV文件已保存: {csv_path}")
    
    # ========================================
    # 生成基础可读文本文件（从CSV数据）
    # ========================================
    try:
        readable_text_path = csv_path.replace('.csv', '_readable.txt')
        with open(readable_text_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write(f"样本分析数据 - 可读文本格式（噪声率: {noise_pct}%）\n")
            f.write("=" * 80 + "\n\n")
            
            # 总览统计
            f.write("\n" + "=" * 80 + "\n")
            f.write("总览统计\n")
            f.write("=" * 80 + "\n\n")
            
            tier_summary = []
            for tier in sorted(df['Tier分级'].unique()):
                tier_df = df[df['Tier分级'] == tier]
                valid_df = tier_df[tier_df['矫正是否正确'] != '不适用(已丢弃)']
                correct_count = len(tier_df[tier_df['矫正是否正确'] == '正确'])
                error_count = len(tier_df[tier_df['矫正是否正确'] == '错误'])
                dropped_count = len(tier_df[tier_df['矫正是否正确'] == '不适用(已丢弃)'])
                
                tier_summary.append({
                    'Tier': tier,
                    '总样本数': len(tier_df),
                    '✓正确处理': correct_count,
                    '✗错误处理': error_count,
                    '⊗已丢弃': dropped_count,
                    '准确率': f"{correct_count / len(valid_df) * 100:.2f}%" if len(valid_df) > 0 else 'N/A',
                    '平均权值': f"{tier_df['样本权值'].mean():.4f}",
                    '噪声样本数': len(tier_df[tier_df['是否噪声'] == '是']),
                    '干净样本数': len(tier_df[tier_df['是否噪声'] == '否'])
                })
            
            summary_df = pd.DataFrame(tier_summary)
            f.write(summary_df.to_string(index=False))
            f.write("\n\n")
            
            # 错误样本汇总
            error_samples = df[df['矫正是否正确'] == '错误']
            if len(error_samples) > 0:
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"错误样本汇总（共{len(error_samples)}个）\n")
                f.write("=" * 80 + "\n\n")
                
                error_key_cols = ['样本ID', 'Tier分级', '是否噪声', '真实标签', '噪声标签', 
                                 '矫正后标签', '矫正动作', '错误类型', '系统置信度']
                error_display_cols = [col for col in error_key_cols if col in error_samples.columns]
                if error_display_cols:
                    f.write(error_samples[error_display_cols].to_string(index=False))
                    f.write("\n\n")
            
            # 误杀样本
            false_positive = df[(df['是否噪声'] == '否') & (df['矫正是否正确'] == '错误')]
            if len(false_positive) > 0:
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"误杀干净样本（共{len(false_positive)}个）\n")
                f.write("=" * 80 + "\n\n")
                
                fp_key_cols = ['样本ID', 'Tier分级', '真实标签', '噪声标签', '矫正后标签', 
                              '矫正动作', '错误类型', '系统置信度']
                fp_display_cols = [col for col in fp_key_cols if col in false_positive.columns]
                if fp_display_cols:
                    f.write(false_positive[fp_display_cols].to_string(index=False))
                    f.write("\n\n")
            
            # 漏网噪声
            false_negative = df[(df['是否噪声'] == '是') & (df['矫正是否正确'] == '错误')]
            if len(false_negative) > 0:
                f.write("\n" + "=" * 80 + "\n")
                f.write(f"漏网噪声（共{len(false_negative)}个）\n")
                f.write("=" * 80 + "\n\n")
                
                fn_key_cols = ['样本ID', 'Tier分级', '真实标签', '噪声标签', '矫正后标签', 
                              '矫正动作', '错误类型', '系统置信度']
                fn_display_cols = [col for col in fn_key_cols if col in false_negative.columns]
                if fn_display_cols:
                    f.write(false_negative[fn_display_cols].to_string(index=False))
                    f.write("\n\n")
            
            # 按Tier分组的样本（前20个）
            f.write("\n" + "=" * 80 + "\n")
            f.write("各Tier样本详情（每个Tier显示前20个样本）\n")
            f.write("=" * 80 + "\n\n")
            
            tier_order = ['P1: Keep', 'P1: Flip', 'P1: Rescue', 'P1: Reweight',
                         'P2: Keep', 'P2: Flip', 'P2: Rescue', 'P2: Reweight']
            
            for tier in tier_order:
                tier_data = df[df['Tier分级'] == tier]
                if len(tier_data) > 0:
                    f.write(f"\n--- {tier} (共{len(tier_data)}个样本) ---\n")
                    
                    key_cols = ['样本ID', '矫正是否正确', '是否噪声', '真实标签', '噪声标签', 
                               '矫正后标签', '矫正动作', '系统置信度', '样本权值']
                    display_cols = [col for col in key_cols if col in tier_data.columns]
                    
                    display_data = tier_data[display_cols].head(20)
                    f.write(display_data.to_string(index=False))
                    if len(tier_data) > 20:
                        f.write(f"\n... (还有{len(tier_data) - 20}个样本未显示)")
                    f.write("\n\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("文件结束\n")
            f.write("=" * 80 + "\n")
        
        logger.info(f"✓ 可读文本文件已保存: {readable_text_path}")
        logger.info(f"  说明: 此文件包含关键统计和样本信息，方便AI直接分析")
    except Exception as e:
        logger.warning(f"⚠ 可读文本文件生成失败: {e}")
        import traceback
        logger.warning(traceback.format_exc())
    
    # Save to Excel with multiple sheets (如果openpyxl可用)
    if EXCEL_AVAILABLE:
        # Excel文件名从CSV文件名生成，已包含噪声率
        excel_path = csv_path.replace('.csv', '.xlsx')
        try:
            from openpyxl.styles import PatternFill, Font
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 重新排列列顺序，把关键信息放前面
                key_columns = [
                    '样本ID', '矫正是否正确', '错误类型', 'Tier分级', 'Tier阶段',
                    '是否噪声', '真实标签', '真实标签值', '噪声标签', '噪声标签值',
                    '矫正后标签', '矫正后标签值', '矫正动作', '决策理由',
                    'Phase2_Action',
                    '系统置信度', '样本权值',
                    'CL疑似噪声', 'CL当前标签置信度', 'CL目标标签置信度',
                    'AUM分数',
                    'KNN邻居标签', 'KNN一致性', 'KNN是否支持当前标签',
                    'iter_CL当前标签置信度', 'iter_CL目标标签置信度', 'iter_CL置信度差值',
                    'stage2_CL疑似噪声',
                    'stage2_AUM分数',
                    'stage2_KNN邻居标签', 'stage2_KNN一致性', 'stage2_KNN是否支持当前标签',
                    '是否翻转', '是否丢弃', '标签是否改变'
                ]
                # 只选择存在的列
                existing_key_columns = [col for col in key_columns if col in df.columns]
                df_reordered = df[existing_key_columns].copy()
                
                # Sheet 1: 总览统计
                tier_summary = []
                for tier in sorted(df['Tier分级'].unique()):
                    tier_df = df[df['Tier分级'] == tier]
                    valid_df = tier_df[tier_df['矫正是否正确'] != '不适用(已丢弃)']
                    correct_count = len(tier_df[tier_df['矫正是否正确'] == '正确'])
                    error_count = len(tier_df[tier_df['矫正是否正确'] == '错误'])
                    dropped_count = len(tier_df[tier_df['矫正是否正确'] == '不适用(已丢弃)'])
                    
                    tier_summary.append({
                        'Tier': tier,
                        '总样本数': len(tier_df),
                        '✓正确处理': correct_count,
                        '✗错误处理': error_count,
                        '⊗已丢弃': dropped_count,
                        '准确率': f"{correct_count / len(valid_df) * 100:.2f}%" if len(valid_df) > 0 else 'N/A',
                        '平均权值': f"{tier_df['样本权值'].mean():.4f}",
                        '噪声样本数': len(tier_df[tier_df['是否噪声'] == '是']),
                        '干净样本数': len(tier_df[tier_df['是否噪声'] == '否'])
                    })
                
                summary_df = pd.DataFrame(tier_summary)
                summary_df.to_excel(writer, sheet_name='📊总览统计', index=False)
                
                # 为总览统计添加格式
                ws = writer.sheets['📊总览统计']
                for row in range(2, len(summary_df) + 2):
                    # 正确处理 - 绿色
                    ws.cell(row=row, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    # 错误处理 - 红色
                    ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    # 已丢弃 - 灰色
                    ws.cell(row=row, column=5).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                # ========================================
                # Sheet: 所有样本统计（阶段1和阶段2）
                # ========================================
                all_samples_stats = []
                
                # 计算Phase1矫正后的标签（用于阶段1统计）
                phase1_labels = []
                for i in range(len(df_reordered)):
                    noisy_label_val = df_reordered.iloc[i]['噪声标签值']
                    phase2_action = df_reordered.iloc[i].get('Phase2_Action', '')
                    final_label_val = df_reordered.iloc[i]['矫正后标签值']
                    
                    # 从phase2_action反推phase1的标签
                    if phase2_action == 'UndoFlip':
                        phase1_label = 1 - noisy_label_val
                    elif phase2_action == 'LateFlip':
                        phase1_label = noisy_label_val
                    else:
                        if final_label_val != noisy_label_val:
                            phase1_label = final_label_val
                        else:
                            phase1_label = noisy_label_val
                    phase1_labels.append(phase1_label)
                
                df_reordered['Phase1矫正后标签值'] = phase1_labels
                df_reordered['Phase1是否正确'] = (df_reordered['Phase1矫正后标签值'] == df_reordered['真实标签值']).apply(lambda x: '正确' if x else '错误')
                
                # 阶段1统计
                total_samples = len(df_reordered)
                phase1_correct = len(df_reordered[df_reordered['Phase1是否正确'] == '正确'])
                phase1_error = len(df_reordered[df_reordered['Phase1是否正确'] == '错误'])
                phase1_accuracy = (phase1_correct / total_samples * 100) if total_samples > 0 else 0
                
                # 阶段1动作统计
                phase1_flip = len(df_reordered[df_reordered['矫正动作'] == 'Flip'])
                phase1_keep = len(df_reordered[df_reordered['矫正动作'] == 'Keep'])
                
                # 阶段1指标统计
                phase1_cl_mean = df_reordered['CL当前标签置信度'].mean() if 'CL当前标签置信度' in df_reordered.columns else 0
                phase1_aum_mean = df_reordered['AUM分数'].mean() if 'AUM分数' in df_reordered.columns else 0
                phase1_knn_mean = df_reordered['KNN一致性'].mean() if 'KNN一致性' in df_reordered.columns else 0
                
                # 阶段2统计（如果有Phase2）
                if 'Phase2_Action' in df_reordered.columns and df_reordered['Phase2_Action'].notna().any():
                    phase2_correct = len(df_reordered[df_reordered['矫正是否正确'] == '正确'])
                    phase2_error = len(df_reordered[df_reordered['矫正是否正确'] == '错误'])
                    phase2_accuracy = (phase2_correct / total_samples * 100) if total_samples > 0 else 0
                    
                    # 阶段2动作统计
                    phase2_undo = len(df_reordered[df_reordered['Phase2_Action'] == 'UndoFlip'])
                    phase2_late = len(df_reordered[df_reordered['Phase2_Action'] == 'LateFlip'])
                    phase2_nochange = len(df_reordered[df_reordered['Phase2_Action'] == 'NoChange'])
                    
                    # 阶段2指标统计
                    phase2_cl_mean = df_reordered['stage2_CL当前标签置信度'].mean() if 'stage2_CL当前标签置信度' in df_reordered.columns else 0
                    phase2_aum_mean = df_reordered['stage2_AUM分数'].mean() if 'stage2_AUM分数' in df_reordered.columns else 0
                    phase2_knn_mean = df_reordered['stage2_KNN一致性'].mean() if 'stage2_KNN一致性' in df_reordered.columns else 0
                else:
                    phase2_correct = 0
                    phase2_error = 0
                    phase2_accuracy = 0
                    phase2_undo = 0
                    phase2_late = 0
                    phase2_nochange = total_samples
                    phase2_cl_mean = 0
                    phase2_aum_mean = 0
                    phase2_knn_mean = 0
                
                # 原始标签统计
                original_correct = len(df_reordered[df_reordered['是否噪声'] == '否'])
                original_error = len(df_reordered[df_reordered['是否噪声'] == '是'])
                original_accuracy = (original_correct / total_samples * 100) if total_samples > 0 else 0
                
                # 构建统计表
                all_samples_stats.append({
                    '统计项': '总样本数',
                    '数值': total_samples,
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '原始标签准确率',
                    '数值': f"{original_accuracy:.2f}%",
                    '阶段1': f"正确: {original_correct}, 错误: {original_error}",
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段1准确率',
                    '数值': f"{phase1_accuracy:.2f}%",
                    '阶段1': f"正确: {phase1_correct}, 错误: {phase1_error}",
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2准确率',
                    '数值': f"{phase2_accuracy:.2f}%" if phase2_accuracy > 0 else 'N/A',
                    '阶段1': '',
                    '阶段2': f"正确: {phase2_correct}, 错误: {phase2_error}" if phase2_accuracy > 0 else 'N/A'
                })
                all_samples_stats.append({
                    '统计项': '阶段1动作',
                    '数值': '',
                    '阶段1': f"Flip: {phase1_flip}, Keep: {phase1_keep}",
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2动作',
                    '数值': '',
                    '阶段1': '',
                    '阶段2': f"UndoFlip: {phase2_undo}, LateFlip: {phase2_late}, NoChange: {phase2_nochange}" if phase2_accuracy > 0 else 'N/A'
                })
                all_samples_stats.append({
                    '统计项': '阶段1平均CL置信度',
                    '数值': f"{phase1_cl_mean:.4f}",
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段1平均AUM分数',
                    '数值': f"{phase1_aum_mean:.4f}",
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段1平均KNN一致性',
                    '数值': f"{phase1_knn_mean:.4f}",
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2平均CL置信度',
                    '数值': f"{phase2_cl_mean:.4f}" if phase2_cl_mean > 0 else 'N/A',
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2平均AUM分数',
                    '数值': f"{phase2_aum_mean:.4f}" if phase2_aum_mean > 0 else 'N/A',
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2平均KNN一致性',
                    '数值': f"{phase2_knn_mean:.4f}" if phase2_knn_mean > 0 else 'N/A',
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段1提升',
                    '数值': f"{phase1_accuracy - original_accuracy:+.2f}%",
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '阶段2提升',
                    '数值': f"{phase2_accuracy - phase1_accuracy:+.2f}%" if phase2_accuracy > 0 else 'N/A',
                    '阶段1': '',
                    '阶段2': ''
                })
                all_samples_stats.append({
                    '统计项': '总提升',
                    '数值': f"{phase2_accuracy - original_accuracy:+.2f}%" if phase2_accuracy > 0 else f"{phase1_accuracy - original_accuracy:+.2f}%",
                    '阶段1': '',
                    '阶段2': ''
                })
                
                stats_df = pd.DataFrame(all_samples_stats)
                stats_df.to_excel(writer, sheet_name='📊所有样本统计', index=False)
                
                # 为统计表添加格式
                ws_stats = writer.sheets['📊所有样本统计']
                # 设置标题行格式
                for col in range(1, len(stats_df.columns) + 1):
                    ws_stats.cell(row=1, column=col).font = Font(bold=True)
                    ws_stats.cell(row=1, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    ws_stats.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
                
                logger.info(f"  ✓ 所有样本统计: 已生成统计表")
                
                # ========================================
                # Sheet: 阶段1指标（所有样本）
                # ========================================
                
                # Phase1矫正后的标签已经在统计部分计算过了，这里直接使用
                if 'Phase1矫正后标签值' not in df_reordered.columns:
                    # 如果还没有计算，则计算一次
                    phase1_labels = []
                    for i in range(len(df_reordered)):
                        noisy_label_val = df_reordered.iloc[i]['噪声标签值']
                        phase2_action = df_reordered.iloc[i].get('Phase2_Action', '')
                        final_label_val = df_reordered.iloc[i]['矫正后标签值']
                        
                        # 从phase2_action反推phase1的标签
                        if phase2_action == 'UndoFlip':
                            phase1_label = 1 - noisy_label_val
                        elif phase2_action == 'LateFlip':
                            phase1_label = noisy_label_val
                        else:
                            if final_label_val != noisy_label_val:
                                phase1_label = final_label_val
                            else:
                                phase1_label = noisy_label_val
                        phase1_labels.append(phase1_label)
                    df_reordered['Phase1矫正后标签值'] = phase1_labels
                
                if 'Phase1矫正后标签' not in df_reordered.columns:
                    df_reordered['Phase1矫正后标签'] = df_reordered['Phase1矫正后标签值'].apply(lambda x: '正常' if x == 0 else '恶意')
                
                # 确保Phase1是否正确列已计算
                if 'Phase1是否正确' not in df_reordered.columns:
                    df_reordered['Phase1是否正确'] = (df_reordered['Phase1矫正后标签值'] == df_reordered['真实标签值']).apply(lambda x: '正确' if x else '错误')
                
                # Sheet: 阶段1指标
                phase1_columns = [
                    '样本ID', '是否噪声', '真实标签', '真实标签值', 
                    '噪声标签', '噪声标签值',
                    'Phase1矫正后标签', 'Phase1矫正后标签值',
                    'CL疑似噪声', 'CL当前标签置信度', 'CL目标标签置信度', 'CL预测标签',
                    'AUM分数',
                    'KNN邻居标签', 'KNN一致性', 'KNN是否支持当前标签',
                    '矫正动作', '系统置信度', '样本权值', '决策理由'
                ]
                phase1_existing_cols = [col for col in phase1_columns if col in df_reordered.columns]
                phase1_df = df_reordered[phase1_existing_cols].copy()
                
                # 添加Phase1是否正确列（如果还没有）
                if 'Phase1是否正确' not in phase1_df.columns:
                    phase1_df['Phase1是否正确'] = (phase1_df['Phase1矫正后标签值'] == phase1_df['真实标签值']).apply(lambda x: '正确' if x else '错误')
                
                # 重新排列列顺序，只选择存在的列
                phase1_column_order = [
                    '样本ID', '是否噪声', '真实标签', '真实标签值', 
                    '噪声标签', '噪声标签值',
                    'Phase1矫正后标签', 'Phase1矫正后标签值', 'Phase1是否正确',
                    'CL疑似噪声', 'CL当前标签置信度', 'CL目标标签置信度', 'CL预测标签',
                    'AUM分数',
                    'KNN邻居标签', 'KNN一致性', 'KNN是否支持当前标签',
                    '矫正动作', '系统置信度', '样本权值', '决策理由'
                ]
                # 只选择存在的列，保持顺序
                phase1_final_cols = [col for col in phase1_column_order if col in phase1_df.columns]
                phase1_df = phase1_df[phase1_final_cols]
                
                phase1_df.to_excel(writer, sheet_name='📊阶段1指标', index=False)
                ws_phase1 = writer.sheets['📊阶段1指标']
                
                # 为阶段1添加颜色标记
                for row in range(2, len(phase1_df) + 2):
                    phase1_correct = phase1_df.iloc[row-2]['Phase1是否正确']
                    if phase1_correct == '正确':
                        for col in range(1, len(phase1_df.columns) + 1):
                            ws_phase1.cell(row=row, column=col).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    elif phase1_correct == '错误':
                        for col in range(1, len(phase1_df.columns) + 1):
                            ws_phase1.cell(row=row, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                
                logger.info(f"  ✓ 阶段1指标: {len(phase1_df)}个样本（所有样本）")
                
                # ========================================
                # Sheet: 阶段2指标（所有样本，基于Phase1矫正后标签重新训练的模型）
                # ========================================
                phase2_columns = [
                    '样本ID', '是否噪声', '真实标签', '真实标签值',
                    '噪声标签', '噪声标签值',  # 添加噪声标签，方便对比
                    'Phase1矫正后标签', 'Phase1矫正后标签值',  # 阶段2的输入标签
                    '矫正后标签', '矫正后标签值',  # 阶段2的输出标签（最终标签）
                    'Phase2_Action',
                    'iter_CL当前标签置信度', 'iter_CL目标标签置信度', 'iter_CL置信度差值',
                    'stage2_CL预测噪声', 'stage2_CL疑似噪声', 'stage2_CL预测标签', 'stage2_CL预测标签值',  # 添加stage2_CL预测噪声（文本格式）和stage2_CL预测标签
                    'stage2_CL当前标签置信度', 'stage2_CL目标标签置信度',  # 添加stage2_CL置信度
                    'stage2_AUM分数',
                    'stage2_KNN邻居标签', 'stage2_KNN邻居标签值', 'stage2_KNN一致性', 'stage2_KNN是否支持当前标签',
                    'Phase2是否正确',  # 添加Phase2是否正确（类似Phase1是否正确）
                    '矫正是否正确', '系统置信度', '样本权值', '决策理由'  # 添加决策理由
                ]
                phase2_existing_cols = [col for col in phase2_columns if col in df_reordered.columns]
                phase2_df = df_reordered[phase2_existing_cols].copy()
                
                # 添加Phase2是否正确列（基于最终矫正后标签）
                if '矫正后标签值' in phase2_df.columns and '真实标签值' in phase2_df.columns:
                    phase2_df['Phase2是否正确'] = (phase2_df['矫正后标签值'] == phase2_df['真实标签值']).apply(lambda x: '正确' if x else '错误')
                
                # 修复阶段2的"KNN是否支持当前标签"：应该基于Phase1矫正后的标签判断
                if 'stage2_KNN邻居标签值' in phase2_df.columns and 'Phase1矫正后标签值' in phase2_df.columns:
                    phase2_df['stage2_KNN是否支持当前标签'] = (
                        phase2_df['stage2_KNN邻居标签值'] == phase2_df['Phase1矫正后标签值']
                    ).apply(lambda x: '是' if x else '否')
                
                # 将stage2_CL疑似噪声从数值转换为文本格式（'是'/'否'），与阶段1保持一致
                if 'stage2_CL疑似噪声' in phase2_df.columns:
                    phase2_df['stage2_CL预测噪声'] = phase2_df['stage2_CL疑似噪声'].apply(
                        lambda x: '是' if (x == 1 or x == True) else ('否' if (x == 0 or x == False) else 'N/A')
                    )
                
                # 重新排列列顺序，只选择存在的列
                phase2_column_order = [
                    '样本ID', '是否噪声', '真实标签', '真实标签值',
                    '噪声标签', '噪声标签值',
                    'Phase1矫正后标签', 'Phase1矫正后标签值',
                    '矫正后标签', '矫正后标签值',
                    'Phase2_Action', 'Phase2是否正确',
                    'iter_CL当前标签置信度', 'iter_CL目标标签置信度', 'iter_CL置信度差值',
                    'stage2_CL预测噪声', 'stage2_CL疑似噪声', 'stage2_CL预测标签', 'stage2_CL预测标签值',
                    'stage2_CL当前标签置信度', 'stage2_CL目标标签置信度',
                    'stage2_AUM分数',
                    'stage2_KNN邻居标签', 'stage2_KNN邻居标签值', 'stage2_KNN一致性', 'stage2_KNN是否支持当前标签',
                    '矫正是否正确', '系统置信度', '样本权值', '决策理由'
                ]
                phase2_final_cols = [col for col in phase2_column_order if col in phase2_df.columns]
                phase2_df = phase2_df[phase2_final_cols]
                
                # 显示所有样本（阶段2指标基于重新训练的模型）
                if len(phase2_df) > 0:
                    phase2_df.to_excel(writer, sheet_name='📊阶段2指标', index=False)
                    ws_phase2 = writer.sheets['📊阶段2指标']
                    
                    # 为阶段2添加颜色标记（基于Phase2是否正确）
                    for row in range(2, len(phase2_df) + 2):
                        phase2_correct = phase2_df.iloc[row-2].get('Phase2是否正确', '')
                        if phase2_correct == '正确':
                            for col in range(1, len(phase2_df.columns) + 1):
                                ws_phase2.cell(row=row, column=col).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        elif phase2_correct == '错误':
                            for col in range(1, len(phase2_df.columns) + 1):
                                ws_phase2.cell(row=row, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    
                    logger.info(f"  ✓ 阶段2指标: {len(phase2_df)}个样本（所有样本，基于重新训练的模型）")
                else:
                    logger.info(f"  ✓ 阶段2指标: 0个样本（无数据）")
            
            logger.info(f"✓ Excel文件已保存（多sheet，带颜色标记）: {excel_path}")
            logger.info(f"  包含 {len(pd.ExcelFile(excel_path).sheet_names)} 个工作表")
            logger.info(f"  颜色说明: 绿色=正确处理, 红色=错误处理, 灰色=已丢弃")
            
            # ========================================
            # 生成Excel详细可读文本文件（供AI分析使用）
            # ========================================
            try:
                excel_text_output_path = excel_path.replace('.xlsx', '_excel_readable.txt')
                with open(excel_text_output_path, 'w', encoding='utf-8') as f:
                    f.write("=" * 80 + "\n")
                    f.write(f"样本分析数据 - Excel详细内容（噪声率: {noise_pct}%）\n")
                    f.write("=" * 80 + "\n\n")
                    
                    # 读取所有sheet
                    excel_file = pd.ExcelFile(excel_path)
                    
                    for sheet_name in excel_file.sheet_names:
                        f.write("\n" + "=" * 80 + "\n")
                        f.write(f"Sheet: {sheet_name}\n")
                        f.write("=" * 80 + "\n\n")
                        
                        sheet_df = pd.read_excel(excel_path, sheet_name=sheet_name)
                        
                        # 如果是统计类sheet，使用更友好的格式
                        if '统计' in sheet_name or '总览' in sheet_name:
                            f.write(sheet_df.to_string(index=False))
                            f.write("\n\n")
                        else:
                            # 对于数据类sheet，限制显示行数，避免文件过大
                            max_rows = 100
                            if len(sheet_df) > max_rows:
                                f.write(f"（仅显示前{max_rows}行，共{len(sheet_df)}行）\n\n")
                                display_df = sheet_df.head(max_rows)
                            else:
                                display_df = sheet_df
                            
                            # 选择关键列显示
                            key_columns = [
                                '样本ID', '矫正是否正确', '错误类型', 'Tier分级', 
                                '是否噪声', '真实标签', '噪声标签', '矫正后标签',
                                '矫正动作', '系统置信度', '样本权值'
                            ]
                            available_key_cols = [col for col in key_columns if col in display_df.columns]
                            other_cols = [col for col in display_df.columns if col not in available_key_cols]
                            
                            # 优先显示关键列
                            if available_key_cols:
                                display_df_key = display_df[available_key_cols]
                                f.write("【关键信息】\n")
                                f.write(display_df_key.to_string(index=False))
                                f.write("\n\n")
                            
                            # 如果有其他列且行数不多，也显示
                            if len(sheet_df) <= 50 and other_cols:
                                display_df_other = display_df[other_cols]
                                f.write("【详细信息】\n")
                                f.write(display_df_other.to_string(index=False))
                                f.write("\n\n")
                    
                    f.write("\n" + "=" * 80 + "\n")
                    f.write("文件结束\n")
                    f.write("=" * 80 + "\n")
                
                logger.info(f"✓ Excel详细文本文件已保存: {excel_text_output_path}")
                logger.info(f"  说明: 此文件包含Excel所有sheet的内容，方便AI直接分析")
            except Exception as e:
                logger.warning(f"⚠ Excel详细文本文件生成失败: {e}")
                import traceback
                logger.warning(traceback.format_exc())
        except Exception as e:
            logger.warning(f"⚠ Excel文件保存失败: {e}")
            import traceback
            logger.warning(traceback.format_exc())
    else:
        logger.info("  提示: 安装openpyxl可生成多sheet Excel文件 (pip install openpyxl)")
    
    # ========================================
    # 新格式：按策略分组的日志总结
    # ========================================
    
    # 按Tier分组样本
    tier_groups = {}
    for i, row in df.iterrows():
        tier = row['Tier分级']
        if tier not in tier_groups:
            tier_groups[tier] = {'correct': [], 'incorrect': [], 'dropped': []}
        
        if row['矫正是否正确'] == '正确':
            tier_groups[tier]['correct'].append(row)
        elif row['矫正是否正确'] == '错误':
            tier_groups[tier]['incorrect'].append(row)
        else:  # 不适用(已丢弃)
            tier_groups[tier]['dropped'].append(row)
    
    # 定义Tier顺序和角色说明（CL+AUM+KNN 新策略）
    tier_order = [
        'Keep',
        'Flip',
        'Reweight',
        'Drop',
        'Dropped'
    ]
    
    role_map = {
        'Keep': 'Keep：CL与AUM均支持，保持原标签（高权重）',
        'Flip': 'Flip：疑似噪声样本，KNN 强支持翻转（高权重）',
        'Reweight': 'Reweight：不确定样本，保持但降权',
        'Drop': 'Drop：疑似噪声样本且不满足翻转条件（丢弃）',
        'Dropped': 'Dropped：已丢弃样本'
    }
    
    # 策略详细说明
    strategy_details = {
        'Keep': {
            'phase': 'Phase 1: Keep',
            'condition': 'CL≥阈值 且 AUM≥阈值',
            'action': '保持原标签',
            'weight': '1.0'
        },
        'Flip': {
            'phase': 'Phase 2: Flip',
            'condition': '候选噪声( CL<阈值 或 AUM<阈值 ) 且 KNN一致性>阈值 且 KNN投票指向目标类',
            'action': '翻转标签',
            'weight': '0.9'
        },
        'Reweight': {
            'phase': 'Phase 2: Reweight',
            'condition': '候选噪声但不满足翻转条件，保持但降权',
            'action': '保持原标签但降权',
            'weight': '0.5'
        },
        'Drop': {
            'phase': 'Phase 3: Drop',
            'condition': '候选噪声且不满足翻转条件（仅当启用Drop）',
            'action': '丢弃样本',
            'weight': '0.0'
        },
        'Dropped': {
            'phase': 'Phase 3: Drop',
            'condition': '已丢弃样本',
            'action': '丢弃样本',
            'weight': '0.0'
        }
    }
    
    # 将文件后缀改为.log以支持更好的表格格式
    log_path = save_path.replace('.txt', '.log')
    
    # Save to formatted log file (新格式 - 表格形式)
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("="*120 + "\n")
        f.write("MEDAL-Lite 标签矫正分析 - 策略分组日志报告\n")
        f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*120 + "\n\n")
        
        # Summary statistics
        f.write("📊 总体统计\n")
        f.write("-"*120 + "\n")
        f.write(f"总样本数: {n_samples}\n")
        f.write(f"  正常流量:  {(y_true == 0).sum()} ({100*(y_true == 0).sum()/n_samples:.1f}%)\n")
        f.write(f"  恶意流量:  {(y_true == 1).sum()} ({100*(y_true == 1).sum()/n_samples:.1f}%)\n\n")
        
        f.write(f"噪声注入率: {100*noise_mask.sum()/n_samples:.1f}%\n")
        f.write(f"  噪声样本数: {noise_mask.sum()}\n\n")
        
        action_mask = results['action_mask']
        f.write(f"矫正动作统计:\n")
        f.write(f"  保持:     {(action_mask == 0).sum()} ({100*(action_mask == 0).sum()/n_samples:.1f}%)\n")
        f.write(f"  翻转:     {(action_mask == 1).sum()} ({100*(action_mask == 1).sum()/n_samples:.1f}%)\n")
        f.write(f"  丢弃:     {(action_mask == 2).sum()} ({100*(action_mask == 2).sum()/n_samples:.1f}%)\n")
        f.write(f"  重加权:   {(action_mask == 3).sum()} ({100*(action_mask == 3).sum()/n_samples:.1f}%)\n\n")
        
        # Correction accuracy (excluding dropped samples)
        keep_mask = action_mask != 2
        corrected_labels = results['clean_labels'][keep_mask]
        true_labels_kept = y_true[keep_mask]
        correction_accuracy = (corrected_labels == true_labels_kept).mean()
        f.write(f"矫正准确率 (不含丢弃样本): {correction_accuracy*100:.2f}%\n")
        f.write(f"  正确: {(corrected_labels == true_labels_kept).sum()}\n")
        f.write(f"  错误: {(corrected_labels != true_labels_kept).sum()}\n\n")
        
        f.write("="*120 + "\n\n")
        
        # 按策略分组的详细分析
        f.write("📋 策略分组详细分析（按阶段和路径）\n")
        f.write("="*120 + "\n\n")
        
        # 按Phase分组
        phase_groups = {
            'Phase 1: 核心严选': [],
            'Phase 2: 分级挽救': [],
            'Phase 3: 锚点拯救': [],
            '已丢弃': []
        }
        
        for tier in tier_order:
            if tier not in tier_groups or len(tier_groups[tier]['correct']) + len(tier_groups[tier]['incorrect']) + len(tier_groups[tier]['dropped']) == 0:
                continue
            
            if 'Tier 1' in tier:
                phase_groups['Phase 1: 核心严选'].append(tier)
            elif 'Tier 2' in tier or 'Tier 3' in tier or 'Tier 4' in tier:
                phase_groups['Phase 2: 分级挽救'].append(tier)
            elif 'Tier 5' in tier:
                phase_groups['Phase 3: 锚点拯救'].append(tier)
            elif 'Dropped' in tier:
                phase_groups['已丢弃'].append(tier)
        
        # 按Phase输出
        for phase_name, tiers in phase_groups.items():
            if not tiers:
                continue
            
            f.write("\n" + "="*120 + "\n")
            f.write(f"【{phase_name}】\n")
            f.write("="*120 + "\n\n")
            
            for tier in tiers:
                if tier not in tier_groups:
                    continue
                
                correct_samples = tier_groups[tier]['correct']
                incorrect_samples = tier_groups[tier]['incorrect']
                dropped_samples = tier_groups[tier]['dropped']
                total_samples = len(correct_samples) + len(incorrect_samples) + len(dropped_samples)
                
                if total_samples == 0:
                    continue
                
                role = role_map.get(tier, '')
                
                # 计算准确率（不包括丢弃的样本）
                valid_samples = len(correct_samples) + len(incorrect_samples)
                accuracy = len(correct_samples) / valid_samples * 100 if valid_samples > 0 else 0
                
                # 获取该Tier的权重
                if correct_samples:
                    weight = correct_samples[0]['样本权值']
                elif incorrect_samples:
                    weight = incorrect_samples[0]['样本权值']
                elif dropped_samples:
                    weight = dropped_samples[0]['样本权值']
                else:
                    weight = 'N/A'
                
                f.write(f"┌─ 【策略】{tier}\n")
                f.write(f"│  角色定位: {role}\n")
                
                # 输出策略详细说明
                if tier in strategy_details:
                    details = strategy_details[tier]
                    f.write(f"│  决策条件: {details['condition']}\n")
                    f.write(f"│  执行动作: {details['action']}\n")
                    f.write(f"│  样本权重: {details['weight']}\n")
                
                f.write(f"│  样本统计: 总数={total_samples} | 实际权重={weight} | 准确率={accuracy:.1f}%\n")
                f.write(f"│             正确={len(correct_samples)} | 错误={len(incorrect_samples)} | 丢弃={len(dropped_samples)}\n")
                f.write(f"└─" + "-"*116 + "\n\n")
                
                # 正确处理的样本 - 表格形式
                if correct_samples:
                    f.write(f"  ✓ 正确处理的样本 ({len(correct_samples)}个):\n")
                    f.write("  " + "-"*150 + "\n")
                    # 表头 - 添加更多关键指标
                    f.write(f"  {'样本ID':>8s} | {'真实':>4s} | {'噪声':>4s} | {'矫正后':>6s} | "
                           f"{'原CL':>6s} | {'iter_CL':>7s} | {'iter_T':>7s} | "
                           f"{'anchor':>7s} | {'orig_KNN':>8s} | 决策理由\n")
                    f.write("  " + "-"*150 + "\n")
                    
                    # 显示前10个样本
                    for row in correct_samples[:10]:
                        # 原始CL置信度
                        orig_cl = row['CL正常概率'] if row['噪声标签']=='正常' else row['CL恶意概率']
                        # 迭代CL
                        iter_cl_current = row.get('iter_CL当前标签置信度', 'N/A')
                        iter_cl_target = row.get('iter_CL目标标签置信度', 'N/A')
                        # 锚点KNN
                        anchor_cons = row.get('anchor_KNN一致性', 'N/A')
                        # 原始KNN
                        orig_knn = row.get('KNN一致性', 'N/A')
                        
                        reason_short = row['决策理由'][:40] + '...' if len(row['决策理由']) > 40 else row['决策理由']
                        f.write(f"  {row['样本ID']:>8d} | {str(row['真实标签']):>4} | {str(row['噪声标签']):>4} | "
                               f"{str(row['矫正后标签']):>6} | "
                               f"{str(orig_cl)[:6]:>6} | {str(iter_cl_current)[:7]:>7} | {str(iter_cl_target)[:7]:>7} | "
                               f"{str(anchor_cons)[:7]:>7} | {str(orig_knn)[:8]:>8} | {reason_short}\n")
                    
                    if len(correct_samples) > 10:
                        f.write(f"  ... 还有 {len(correct_samples) - 10} 个正确样本（详见Excel文件）\n")
                    f.write("\n")
                
                # 错误处理的样本 - 表格形式（显示所有）
                if incorrect_samples:
                    f.write(f"  ✗ 错误处理的样本 ({len(incorrect_samples)}个):\n")
                    f.write("  " + "-"*150 + "\n")
                    # 表头
                    f.write(f"  {'样本ID':>8s} | {'真实':>4s} | {'噪声':>4s} | {'矫正后':>6s} | "
                           f"{'原CL':>6s} | {'iter_CL':>7s} | {'iter_T':>7s} | "
                           f"{'anchor':>7s} | {'orig_KNN':>8s} | 错误原因\n")
                    f.write("  " + "-"*150 + "\n")
                    
                    for row in incorrect_samples:
                        # 原始CL置信度
                        orig_cl = row['CL正常概率'] if row['噪声标签']=='正常' else row['CL恶意概率']
                        # 迭代CL
                        iter_cl_current = row.get('iter_CL当前标签置信度', 'N/A')
                        iter_cl_target = row.get('iter_CL目标标签置信度', 'N/A')
                        # 锚点KNN
                        anchor_cons = row.get('anchor_KNN一致性', 'N/A')
                        # 原始KNN
                        orig_knn = row.get('KNN一致性', 'N/A')
                        
                        error_reason = f"真实={row['真实标签']}, 矫正为{row['矫正后标签']}"
                        f.write(f"  {row['样本ID']:>8d} | {str(row['真实标签']):>4} | {str(row['噪声标签']):>4} | "
                               f"{str(row['矫正后标签']):>6} | "
                               f"{str(orig_cl)[:6]:>6} | {str(iter_cl_current)[:7]:>7} | {str(iter_cl_target)[:7]:>7} | "
                               f"{str(anchor_cons)[:7]:>7} | {str(orig_knn)[:8]:>8} | {error_reason}\n")
                    f.write("\n")
                
                # 丢弃的样本 - 表格形式
                if dropped_samples:
                    f.write(f"  🗑 丢弃的样本 ({len(dropped_samples)}个):\n")
                    f.write("  " + "-"*150 + "\n")
                    # 表头
                    f.write(f"  {'样本ID':>8s} | {'真实':>4s} | {'噪声':>4s} | {'是否噪声':>8s} | "
                           f"{'原CL':>6s} | {'iter_CL':>7s} | {'iter_T':>7s} | "
                           f"{'anchor':>7s} | {'orig_KNN':>8s} | 丢弃原因\n")
                    f.write("  " + "-"*150 + "\n")
                    
                    # 显示前10个丢弃样本
                    for row in dropped_samples[:10]:
                        # 原始CL置信度
                        orig_cl = row['CL正常概率'] if row['噪声标签']=='正常' else row['CL恶意概率']
                        # 迭代CL
                        iter_cl_current = row.get('iter_CL当前标签置信度', 'N/A')
                        iter_cl_target = row.get('iter_CL目标标签置信度', 'N/A')
                        # 锚点KNN
                        anchor_cons = row.get('anchor_KNN一致性', 'N/A')
                        # 原始KNN
                        orig_knn = row.get('KNN一致性', 'N/A')
                        
                        drop_reason = "CL极低" if 'CL' in row['决策理由'] else "无法挽救"
                        f.write(f"  {row['样本ID']:>8d} | {str(row['真实标签']):>4} | {str(row['噪声标签']):>4} | "
                               f"{str(row['是否噪声']):>8} | "
                               f"{str(orig_cl)[:6]:>6} | {str(iter_cl_current)[:7]:>7} | {str(iter_cl_target)[:7]:>7} | "
                               f"{str(anchor_cons)[:7]:>7} | {str(orig_knn)[:8]:>8} | {drop_reason}\n")
                    
                    if len(dropped_samples) > 10:
                        f.write(f"  ... 还有 {len(dropped_samples) - 10} 个丢弃样本（详见Excel文件）\n")
                    f.write("\n")
                
                f.write("\n")
        
        # 添加误杀分析（被错误丢弃的干净样本）
        f.write("="*120 + "\n")
        f.write("⚠️  误杀分析 - 被错误丢弃的干净样本\n")
        f.write("="*120 + "\n\n")
        
        # 找出所有被丢弃但实际是干净的样本
        false_drops = []
        for i, row in df.iterrows():
            if row['矫正动作'] == '丢弃' and row['是否噪声'] == '否':
                false_drops.append(row)
        
        if false_drops:
            f.write(f"📊 被误杀的干净样本详情 (共{len(false_drops)}个):\n")
            f.write("-"*120 + "\n")
            f.write(f"  {'样本ID':>8s} | {'标签':>4s} | {'iter_CL':>7s} | {'iter_CL_T':>9s} | {'anchor':>7s} | {'orig_KNN':>8s} | 原因\n")
            f.write("-"*120 + "\n")
            
            for row in false_drops:
                # 使用iter_CL当前标签置信度（这才是Drop判断的依据）
                iter_cl_current = row.get('iter_CL当前标签置信度', 'N/A')
                iter_cl_target = row.get('iter_CL目标标签置信度', 'N/A')
                anchor_cons = row.get('anchor_KNN一致性', 'N/A')
                orig_knn_cons = row.get('KNN一致性', 'N/A')
                
                f.write(f"  {row['样本ID']:>8d} | {str(row['真实标签']):>4} | {str(iter_cl_current):>7} | "
                       f"{str(iter_cl_target):>9} | {str(anchor_cons):>7} | {str(orig_knn_cons):>8} | CL极低\n")
        else:
            f.write("✓ 没有误杀干净样本\n")
        
        f.write("\n")
        
        # 添加漏网之鱼分析（未被检测到的噪声样本）
        f.write("="*120 + "\n")
        f.write("🐟 漏网之鱼分析 - 未被矫正的噪声样本\n")
        f.write("="*120 + "\n\n")
        
        # 找出所有是噪声但未被翻转的样本
        missed_noise = []
        for i, row in df.iterrows():
            if row['是否噪声'] == '是' and row['矫正动作'] != '翻转':
                missed_noise.append(row)
        
        if missed_noise:
            f.write(f"📊 漏网的噪声样本详情 (共{len(missed_noise)}个):\n")
            f.write("-"*120 + "\n")
            f.write(f"  {'样本ID':>8s} | {'真实':>4s} | {'噪声':>4s} | {'动作':>4s} | "
                   f"{'CL置信':>7s} | {'AUM':>7s} | {'KNN一致':>7s} | 原因\n")
            f.write("-"*120 + "\n")
            
            for row in missed_noise[:20]:  # 显示前20个
                cl_conf = row['CL正常概率'] if row['噪声标签']=='正常' else row['CL恶意概率']
                reason = "未被CL检测" if row['CL疑似噪声'] == '否' else "KNN支持错误标签"
                aum_val = row.get('AUM分数', 'N/A')
                f.write(f"  {row['样本ID']:>8d} | {str(row['真实标签']):>4} | {str(row['噪声标签']):>4} | "
                       f"{str(row['矫正动作']):>4} | {str(cl_conf):>7} | {str(aum_val)[:7]:>7} | "
                       f"{str(row['KNN一致性']):>7} | {reason}\n")
            
            if len(missed_noise) > 20:
                f.write(f"  ... 还有 {len(missed_noise) - 20} 个漏网样本（详见CSV文件）\n")
        else:
            f.write("✓ 所有噪声样本都被成功检测\n")
        
        f.write("\n")
        f.write("="*120 + "\n")
        f.write("报告结束\n")
        f.write("="*120 + "\n")
    
    logger.info(f"✓ 策略分组日志已保存: {log_path}")
    logger.info(f"  已分析样本总数: {n_samples}")
    logger.info(f"  策略分组数: {len([t for t in tier_order if t in tier_groups])}")
    
    return df


def plot_feature_distributions(results, y_true, noise_mask, save_dir, logger):
    """
    Generate feature distribution plots for clean, noisy, and corrected data
    
    Args:
        results: dict from run_hybrid_court_with_detailed_tracking
        y_true: (N,) ground truth labels
        noise_mask: (N,) boolean array indicating noise
        save_dir: directory to save plots
        logger: logger
    """
    logger.info("")
    logger.info("="*70)
    logger.info("Generating Feature Distribution Plots")
    logger.info("="*70)
    
    features = results['features']
    noisy_labels = results['noisy_labels']
    clean_labels = results['clean_labels']
    action_mask = results['action_mask']
    
    os.makedirs(save_dir, exist_ok=True)
    
    # Use t-SNE for dimensionality reduction
    logger.info("Running t-SNE for dimensionality reduction...")
    tsne = TSNE(n_components=2, random_state=42, perplexity=30, max_iter=1000)
    features_2d = tsne.fit_transform(features)
    logger.info("✓ t-SNE complete")
    
    # Define colors
    color_benign = '#2E86AB'  # Blue
    color_malicious = '#A23B72'  # Red
    
    # 1. Ground Truth Labels (Clean Data)
    logger.info("绘制 1/4: 真实标签分布...")
    fig, ax = plt.subplots(figsize=(10, 8))
    
    benign_mask = y_true == 0
    malicious_mask = y_true == 1
    
    ax.scatter(features_2d[benign_mask, 0], features_2d[benign_mask, 1],
               c=color_benign, label=f'正常流量 (n={benign_mask.sum()})', 
               alpha=0.6, s=30, edgecolors='k', linewidths=0.3)
    ax.scatter(features_2d[malicious_mask, 0], features_2d[malicious_mask, 1],
               c=color_malicious, label=f'恶意流量 (n={malicious_mask.sum()})', 
               alpha=0.6, s=30, edgecolors='k', linewidths=0.3)
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title('特征分布: 真实标签 (干净数据)', 
                fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    
    save_path = os.path.join(save_dir, '1_真实标签分布.png')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    logger.info(f"  ✓ 已保存: {save_path}")
    
    # 2. Noisy Labels (with noise highlighted)
    logger.info("绘制 2/4: 噪声标签分布...")
    fig, ax = plt.subplots(figsize=(10, 8))
    
    # Plot clean samples first
    clean_samples = ~noise_mask
    benign_clean = clean_samples & (noisy_labels == 0)
    malicious_clean = clean_samples & (noisy_labels == 1)
    
    ax.scatter(features_2d[benign_clean, 0], features_2d[benign_clean, 1],
               c=color_benign, label=f'正常流量 (干净, n={benign_clean.sum()})', 
               alpha=0.4, s=25, edgecolors='k', linewidths=0.3)
    ax.scatter(features_2d[malicious_clean, 0], features_2d[malicious_clean, 1],
               c=color_malicious, label=f'恶意流量 (干净, n={malicious_clean.sum()})', 
               alpha=0.4, s=25, edgecolors='k', linewidths=0.3)
    
    # Plot noisy samples with special marker
    noisy_samples = noise_mask
    benign_noisy = noisy_samples & (noisy_labels == 0)
    malicious_noisy = noisy_samples & (noisy_labels == 1)
    
    ax.scatter(features_2d[benign_noisy, 0], features_2d[benign_noisy, 1],
               c=color_benign, label=f'正常流量 (噪声, n={benign_noisy.sum()})', 
               alpha=0.9, s=80, edgecolors='yellow', linewidths=2, marker='s')
    ax.scatter(features_2d[malicious_noisy, 0], features_2d[malicious_noisy, 1],
               c=color_malicious, label=f'恶意流量 (噪声, n={malicious_noisy.sum()})', 
               alpha=0.9, s=80, edgecolors='yellow', linewidths=2, marker='s')
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title(f'特征分布: 噪声标签 (噪声率: {100*noise_mask.sum()/len(noise_mask):.1f}%)', 
                fontsize=14, fontweight='bold')
    ax.legend(fontsize=9, loc='best')
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    
    save_path = os.path.join(save_dir, '2_噪声标签分布.png')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    logger.info(f"  ✓ 已保存: {save_path}")
    
    # 3. Corrected Labels (excluding dropped samples)
    logger.info("绘制 3/4: 矫正后标签分布...")
    fig, ax = plt.subplots(figsize=(10, 8))
    
    kept_samples = action_mask != 2  # Exclude dropped samples
    
    benign_corrected = kept_samples & (clean_labels == 0)
    malicious_corrected = kept_samples & (clean_labels == 1)
    dropped_samples = action_mask == 2
    
    ax.scatter(features_2d[benign_corrected, 0], features_2d[benign_corrected, 1],
               c=color_benign, label=f'正常流量 (n={benign_corrected.sum()})', 
               alpha=0.6, s=30, edgecolors='k', linewidths=0.3)
    ax.scatter(features_2d[malicious_corrected, 0], features_2d[malicious_corrected, 1],
               c=color_malicious, label=f'恶意流量 (n={malicious_corrected.sum()})', 
               alpha=0.6, s=30, edgecolors='k', linewidths=0.3)
    
    # Show dropped samples
    if dropped_samples.sum() > 0:
        ax.scatter(features_2d[dropped_samples, 0], features_2d[dropped_samples, 1],
                   c='gray', label=f'已丢弃 (n={dropped_samples.sum()})', 
                   alpha=0.5, s=50, marker='x', linewidths=1.5)
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title('特征分布: 矫正后标签 (Hybrid Court矫正后)', 
                fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    
    save_path = os.path.join(save_dir, '3_矫正后标签分布.png')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    logger.info(f"  ✓ 已保存: {save_path}")
    
    # 4. Action Visualization (Keep, Flip, Drop, Reweight)
    logger.info("绘制 4/4: 矫正动作分布...")
    fig, ax = plt.subplots(figsize=(12, 9))
    
    keep_mask = action_mask == 0
    flip_mask = action_mask == 1
    drop_mask = action_mask == 2
    reweight_mask = action_mask == 3
    
    # Plot in reverse order so important actions are on top
    if reweight_mask.sum() > 0:
        ax.scatter(features_2d[reweight_mask, 0], features_2d[reweight_mask, 1],
                   c='#F77F00', label=f'重加权 (长尾样本, n={reweight_mask.sum()})', 
                   alpha=0.7, s=60, edgecolors='k', linewidths=0.5, marker='v')
    
    if keep_mask.sum() > 0:
        ax.scatter(features_2d[keep_mask, 0], features_2d[keep_mask, 1],
                   c='#06A77D', label=f'保持 (核心样本, n={keep_mask.sum()})', 
                   alpha=0.5, s=25, edgecolors='k', linewidths=0.3)
    
    if flip_mask.sum() > 0:
        ax.scatter(features_2d[flip_mask, 0], features_2d[flip_mask, 1],
                   c='#D62828', label=f'翻转 (已矫正, n={flip_mask.sum()})', 
                   alpha=0.8, s=70, edgecolors='k', linewidths=0.8, marker='*')
    
    if drop_mask.sum() > 0:
        ax.scatter(features_2d[drop_mask, 0], features_2d[drop_mask, 1],
                   c='#6A4C93', label=f'丢弃 (系统噪声, n={drop_mask.sum()})', 
                   alpha=0.8, s=80, edgecolors='k', linewidths=1, marker='X')
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title('特征分布: Hybrid Court矫正动作', 
                fontsize=14, fontweight='bold')
    ax.legend(fontsize=11, loc='best', framealpha=0.9)
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    
    save_path = os.path.join(save_dir, '4_矫正动作分布.png')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    logger.info(f"  ✓ 已保存: {save_path}")
    
    # 5. Comparison Plot (Before vs After)
    logger.info("绘制 5/5: 矫正前后对比...")
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    
    # Before: Noisy Labels
    ax = axes[0]
    benign_noisy_all = noisy_labels == 0
    malicious_noisy_all = noisy_labels == 1
    
    ax.scatter(features_2d[benign_noisy_all, 0], features_2d[benign_noisy_all, 1],
               c=color_benign, label=f'正常流量 (n={benign_noisy_all.sum()})', 
               alpha=0.5, s=30, edgecolors='k', linewidths=0.3)
    ax.scatter(features_2d[malicious_noisy_all, 0], features_2d[malicious_noisy_all, 1],
               c=color_malicious, label=f'恶意流量 (n={malicious_noisy_all.sum()})', 
               alpha=0.5, s=30, edgecolors='k', linewidths=0.3)
    
    # Highlight noise
    ax.scatter(features_2d[noise_mask, 0], features_2d[noise_mask, 1],
               facecolors='none', edgecolors='yellow', 
               label=f'噪声样本 (n={noise_mask.sum()})', 
               s=100, linewidths=2, marker='o')
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title(f'矫正前: 噪声标签\n(噪声率: {100*noise_mask.sum()/len(noise_mask):.1f}%)', 
                fontsize=13, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    
    # After: Corrected Labels
    ax = axes[1]
    benign_corrected_all = kept_samples & (clean_labels == 0)
    malicious_corrected_all = kept_samples & (clean_labels == 1)
    
    ax.scatter(features_2d[benign_corrected_all, 0], features_2d[benign_corrected_all, 1],
               c=color_benign, label=f'正常流量 (n={benign_corrected_all.sum()})', 
               alpha=0.5, s=30, edgecolors='k', linewidths=0.3)
    ax.scatter(features_2d[malicious_corrected_all, 0], features_2d[malicious_corrected_all, 1],
               c=color_malicious, label=f'恶意流量 (n={malicious_corrected_all.sum()})', 
               alpha=0.5, s=30, edgecolors='k', linewidths=0.3)
    
    # Show corrections
    ax.scatter(features_2d[flip_mask, 0], features_2d[flip_mask, 1],
               facecolors='none', edgecolors='lime', 
               label=f'已翻转 (n={flip_mask.sum()})', 
               s=100, linewidths=2, marker='o')
    
    if drop_mask.sum() > 0:
        ax.scatter(features_2d[drop_mask, 0], features_2d[drop_mask, 1],
                   c='gray', label=f'已丢弃 (n={drop_mask.sum()})', 
                   alpha=0.6, s=50, marker='x', linewidths=1.5)
    
    # Calculate correction accuracy
    correction_accuracy = (clean_labels[kept_samples] == y_true[kept_samples]).mean()
    
    ax.set_xlabel('t-SNE 第一主成分', fontsize=12)
    ax.set_ylabel('t-SNE 第二主成分', fontsize=12)
    ax.set_title(f'矫正后: Hybrid Court\n(准确率: {correction_accuracy*100:.2f}%)', 
                fontsize=13, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    save_path = os.path.join(save_dir, '5_矫正前后对比.png')
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.close()
    logger.info(f"  ✓ 已保存: {save_path}")
    
    logger.info("="*70)
    logger.info("✓ 所有特征分布图已生成")


def main(args):
    """Main function"""
    global logger
    
    # Setup
    set_seed(config.SEED)
    config.create_dirs()
    
    # 噪声率百分比
    noise_pct = int(args.noise_rate * 100)
    
    # Create analysis output directory (包含噪声率)
    analysis_dir = os.path.join(config.LABEL_CORRECTION_DIR, "analysis", f"noise_{noise_pct}pct")
    os.makedirs(analysis_dir, exist_ok=True)
    os.makedirs(os.path.join(analysis_dir, "figures"), exist_ok=True)
    os.makedirs(os.path.join(analysis_dir, "documents"), exist_ok=True)
    
    # 创建带噪声率的日志文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_filename = f"noise_{noise_pct}pct_analysis_{timestamp}.log"
    log_dir = os.path.join(config.OUTPUT_ROOT, "logs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, log_filename)
    
    # 创建共享的handlers
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')
    
    fh = logging.FileHandler(log_path, encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)
    
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    
    # 配置root logger (捕获所有模块的日志，包括HybridCourt)
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)
    root_logger.handlers = []  # 清除已有handlers
    root_logger.addHandler(fh)
    root_logger.addHandler(ch)
    
    # 获取本模块的logger (使用root logger，不需要单独配置)
    logger = logging.getLogger('label_correction_analysis')
    logger.setLevel(logging.INFO)
    logger.handlers = []  # 清除handlers，使用root的
    logger.propagate = True  # 传播到root logger
    
    # 简洁的标题
    logger.info("")
    logger.info("╔" + "═"*68 + "╗")
    logger.info("║" + f"  MEDAL-Lite 标签矫正分析 - 噪声率 {noise_pct}%".center(68) + "║")
    logger.info("╚" + "═"*68 + "╝")
    logger.info(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  设备: {config.DEVICE}")
    logger.info("")
    
    # Override config with arguments
    if args.noise_rate is not None:
        config.LABEL_NOISE_RATE = args.noise_rate
    
    # ========================
    # 1. Load Dataset
    # ========================
    logger.info("┌─ Step 1: 加载数据集")
    logger.info(f"│  数据路径: {config.BENIGN_TRAIN} | {config.MALICIOUS_TRAIN}")
    logger.info(f"│  序列长度: {config.SEQUENCE_LENGTH}")
    
    # 优先使用预处理好的数据
    if check_preprocessed_exists('train'):
        logger.info("│  ✓ 使用预处理文件")
        X_train, y_train_clean, train_files = load_preprocessed('train')
    else:
        # 从PCAP文件加载
        logger.info("│  ⚠ 从PCAP文件加载（建议先预处理以加速）")
        X_train, y_train_clean, train_files = load_dataset(
            benign_dir=config.BENIGN_TRAIN,
            malicious_dir=config.MALICIOUS_TRAIN,
            sequence_length=config.SEQUENCE_LENGTH
        )
    
    if X_train is None:
        logger.error("│  ❌ 数据集加载失败!")
        return
    
    logger.info(f"└─ ✓ 完成: {X_train.shape[0]} 个样本 | 正常={(y_train_clean==0).sum()} | 恶意={(y_train_clean==1).sum()}")
    logger.info("")
    
    # ========================
    # 2. Inject Label Noise
    # ========================
    logger.info("┌─ Step 2: 注入标签噪声")
    logger.info(f"│  噪声率: {config.LABEL_NOISE_RATE*100:.0f}%")
    
    # 固定随机种子，确保相同噪声率的结果可复现
    set_seed(config.SEED)
    y_train_noisy, noise_mask = inject_label_noise(y_train_clean, config.LABEL_NOISE_RATE)
    
    logger.info(f"└─ ✓ 完成: {noise_mask.sum()} 个标签被翻转 | 原始纯度: {100*(y_train_clean==y_train_noisy).mean():.1f}%")
    logger.info("")
    
    # ========================
    # 3. Extract Features
    # ========================
    logger.info("┌─ Step 3: 提取特征")
    
    # 确定backbone路径：优先使用命令行参数，否则使用默认路径
    if args.backbone_path:
        backbone_path = args.backbone_path
    else:
        backbone_path = os.path.join(config.FEATURE_EXTRACTION_DIR, "models", "backbone_pretrained.pth")
    
    # Try to load pre-trained backbone
    if os.path.exists(backbone_path) and not args.retrain_backbone:
        logger.info(f"│  加载预训练backbone: {os.path.basename(backbone_path)}")
        # 使用安全的模型加载函数（自动处理兼容性）
        from MoudleCode.utils.model_loader import load_backbone_safely
        backbone = load_backbone_safely(
            backbone_path=backbone_path,
            config=config,
            device=config.DEVICE,
            logger=logger
        )
        logger.info("│  ✓ Backbone加载完成")
    else:
        # 构建新模型
        backbone = build_backbone(config)
        if args.retrain_backbone:
            logger.info("│  ⚠ 使用随机初始化的backbone")
        else:
            logger.info(f"│  ⚠ 未找到预训练backbone，使用随机初始化")
        logger.info("│  ✓ Backbone初始化完成")
    
    features = extract_features_with_backbone(backbone, X_train, config, logger)
    
    # Save features
    features_path = os.path.join(analysis_dir, "extracted_features.npy")
    np.save(features_path, features)
    logger.info(f"└─ ✓ 完成: 特征维度 {features.shape} | 已保存")
    logger.info("")
    
    # ========================
    # 4. Run Hybrid Court Label Correction
    # ========================
    logger.info("┌─ Step 4: 运行 Hybrid Court 标签矫正")
    
    results = run_hybrid_court_with_detailed_tracking(features, y_train_noisy, config, logger, y_true=y_train_clean)
    
    # Save results
    results_path = os.path.join(analysis_dir, "correction_results.npz")
    np.savez(results_path,
             y_clean=y_train_clean,
             y_noisy=y_train_noisy,
             y_corrected=results['clean_labels'],
             action_mask=results['action_mask'],
             confidence=results['confidence'],
             correction_weight=results['correction_weight'],
             noise_mask=noise_mask,
             features=features,
             cl_suspected_noise=results['cl_suspected_noise'],
             cl_pred_labels=results['cl_pred_labels'],
             cl_pred_probs=results['cl_pred_probs'],
             aum_scores=results['aum_scores'],
             knn_neighbor_labels=results['knn_neighbor_labels'],
             knn_neighbor_consistency=results['knn_neighbor_consistency'],
             tier_info=np.array(results.get('tier_info', []), dtype=object))
    
    # 计算关键指标
    action_mask = results['action_mask']
    keep_mask = action_mask != 2
    correction_accuracy = (results['clean_labels'][keep_mask] == y_train_clean[keep_mask]).mean()
    original_purity = (y_train_clean == y_train_noisy).mean()
    final_purity = (results['clean_labels'][keep_mask] == y_train_clean[keep_mask]).mean()
    improvement = final_purity - original_purity
    
    logger.info(f"└─ ✓ 完成: 原始纯度 {original_purity*100:.1f}% → 最终纯度 {final_purity*100:.1f}% (提升 +{improvement*100:.1f}%)")
    logger.info("")
    
    # ========================
    # 5. Generate Analysis Document
    # ========================
    logger.info("┌─ Step 5: 生成样本分析文档")
    doc_path = os.path.join(analysis_dir, "documents", f"sample_analysis_{noise_pct}pct_cl_aum_knn.log")
    df_analysis = generate_sample_analysis_document(
        results, y_train_clean, noise_mask, doc_path, logger, noise_pct=noise_pct
    )
    logger.info(f"└─ ✓ 完成: CSV/Excel/LOG 已生成")
    logger.info("")
    
    # ========================
    # 6. Generate Feature Distribution Plots
    # ========================
    logger.info("┌─ Step 6: 生成特征分布图")
    plot_dir = os.path.join(analysis_dir, "figures")
    plot_feature_distributions(results, y_train_clean, noise_mask, plot_dir, logger)
    logger.info(f"└─ ✓ 完成: 5 张图表已生成")
    logger.info("")
    
    # ========================
    # Summary
    # ========================
    logger.info("╔" + "═"*68 + "╗")
    logger.info("║" + f"  ✓ 噪声率 {noise_pct}% 分析完成".center(68) + "║")
    logger.info("╚" + "═"*68 + "╝")
    logger.info("")
    
    # 关键统计摘要（紧凑格式）
    n_samples = len(y_train_clean)
    action_mask = results['action_mask']
    keep_mask = action_mask != 2
    correction_accuracy = (results['clean_labels'][keep_mask] == y_train_clean[keep_mask]).mean()
    original_purity = (y_train_clean == y_train_noisy).mean()
    final_purity = correction_accuracy
    improvement = final_purity - original_purity
    
    logger.info("📊 关键指标:")
    logger.info(f"   原始纯度: {original_purity*100:.1f}%  →  最终纯度: {final_purity*100:.1f}%  (提升: +{improvement*100:.1f}%)")
    logger.info(f"   动作分布: 未翻转 {(action_mask==0).sum()} ({100*(action_mask==0).sum()/n_samples:.1f}%) | "
                f"翻转 {(action_mask==1).sum()} ({100*(action_mask==1).sum()/n_samples:.1f}%) | "
                f"丢弃 {(action_mask==2).sum()} ({100*(action_mask==2).sum()/n_samples:.1f}%)")
    logger.info("")
    
    # 输出可解析的摘要行（供Shell脚本提取）
    logger.info(f"SUMMARY: noise_rate={noise_pct}% | original_purity={original_purity*100:.2f}% | "
                f"final_purity={final_purity*100:.2f}% | improvement={improvement*100:.2f}% | "
                f"flip_count={(action_mask==1).sum()} | keep_count={(action_mask==0).sum()}")
    logger.info("")
    
    # 清理root logger的handlers
    for handler in root_logger.handlers[:]:
        handler.close()
        root_logger.removeHandler(handler)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Label Correction Analysis for MEDAL-Lite"
    )
    parser.add_argument(
        "--noise_rate", 
        type=float, 
        default=0.30, 
        help="Label noise rate (default: 0.30)"
    )
    parser.add_argument(
        "--retrain_backbone",
        action='store_true',
        help="Use randomly initialized backbone instead of pre-trained"
    )
    parser.add_argument(
        "--backbone_path",
        type=str,
        default='',
        help="Path to backbone model (optional, default: backbone_pretrained.pth)"
    )
    
    args = parser.parse_args()
    
    main(args)

